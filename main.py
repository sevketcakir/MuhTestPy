import sys
import os
import logging
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem
from PySide6.QtCore import QObject, Signal, Slot, QSettings
from uiloader import loadUi
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill
from pathlib import Path
from learning_outcome import LearningOutcomeModel, ComboBoxDelegate
from exam import Ogrenci, Exam, Answers

# ------ Logging --------
class LogEmitter(QObject):
    """A QObject that emits signals for logging."""
    log_signal = Signal(str)

class QLogHandler(logging.Handler):
    """
    A custom logging handler that emits signals to be connected 
    to a QPlainTextEdit widget.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.emitter = LogEmitter()

    def emit(self, record):
        """
        This method is called by the logging module when a log message
        is generated. We format the message and emit it.
        """
        msg = self.format(record)
        self.emitter.log_signal.emit(msg)
# ------ Logging --------


class MyMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi("mainwindow.ui", self)
        self.dosyaAc.clicked.connect(self.dosya_ac)
        self.exceleAktar.clicked.connect(self.excele_aktar)
        self.okIliskilendir.stateChanged.connect(self.iliskilendir)
        self.table_model = None
        self.exam = None
        self.settings = QSettings("PAUMuhendislik", "SinavOkuma")
        self.last_directory = self.settings.value("last_directory", os.getcwd())
        self.set_table_headers()
        if hasattr(self, 'bilgiMesaji'):
                    self.bilgiMesaji.setReadOnly(True)
        else:
            print("WARNING: 'bilgiMesaji' widget not found in .ui file. Logging will only go to console.")
        
        self.setup_logging()
        logging.info("Uygulama başladı. Sınav dosyası yüklenmeye hazır.")

    @Slot(str) 
    def append_log_message(self, message):
        """Append a log message to the bilgiMesaji widget."""
        if hasattr(self, 'bilgiMesaji'):
            self.bilgiMesaji.appendPlainText(message)

    def setup_logging(self):
        """Configure the logging system."""
        
        # 1. Create our custom QLogHandler
        self.log_handler = QLogHandler()

        # 2. Create a formatter
        formatter = logging.Formatter(
            '%(asctime)s [%(levelname)-8s] (%(module)s) %(message)s',
            datefmt='%H:%M:%S'
        )
        self.log_handler.setFormatter(formatter)

        # 3. Connect the handler's signal to our slot
        self.log_handler.emitter.log_signal.connect(self.append_log_message)

        # 4. Get the root logger and set its level
        logger = logging.getLogger()
        logger.setLevel(logging.INFO) # Log INFO and above (WARNING, ERROR)
        
        # 5. Add our custom handler to the root logger
        logger.addHandler(self.log_handler)

        # (Optional) Add a standard console logger too
        stdout_handler = logging.StreamHandler(sys.stdout)
        stdout_handler.setFormatter(formatter)
        logger.addHandler(stdout_handler)

    def iliskilendir(self):
        #print(self.okIliskilendir.isChecked())
        if self.okIliskilendir.isChecked():
            soru_sayisi = self.exam.answers[max(self.exam.answers, key=lambda a:self.exam.answers[a].question_count)].question_count
            qlist = [f"Soru {i+1}" for i in range(soru_sayisi)]
            oksayisi = self.okSayisi.value()
            oklist = [f"ÖK{i+1}" for i in range(oksayisi)]
            gruplar = list(self.exam.answers.keys())
            self.table_model = LearningOutcomeModel(qlist, oklist, gruplar)
            self.okTablo.setModel(self.table_model)

            delegate = ComboBoxDelegate(oklist)
            for sutun in range(1, len(gruplar)+1):
                self.okTablo.setItemDelegateForColumn(sutun, delegate)



    def get_excel_column_name(self):
        s1 = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        s2 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for a in s1:
            for b in s2:
                yield (a+b).strip()

    def default_sheet(self, ws, style):
        ws.title = "Notlar"
        for w,c in zip([90, 300, 50, 50, 50, 50, 50, 50],"ABCDEFGH"):
            ws.column_dimensions[c].width = w*0.13

        ws.append(["Pamukkale Üniversitesi Mühendislik Fakültesi Test Sonuçları"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.cell(1,1).style = style
        ws.append([f"Dersin adı: {self.dersinAdi.text()}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.cell(2,1).style = style
        ws.append([])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
        ws.append(["Öğr.No.", "Adı", "Şube", "Kitapçık Türü", "Doğru", "Yanlış", "Boş", "Puan"])

        if self.ogrnoSirala.isChecked():
            ogrenciler = sorted(self.exam.students, key=lambda ogr: ogr.ogrno)
        elif self.notSirala.isChecked():
            ogrenciler = sorted(self.exam.students, key=lambda ogr: ogr.puan, reverse=True)
        else:
            ogrenciler = self.exam.students

        for ogrenci in ogrenciler:
            ws.append([ogrenci.ogrno,
                       ogrenci.adi,
                       ogrenci.sube,
                       ogrenci.kitapcik,
                       ogrenci.dogru,
                       ogrenci.yanlis,
                       ogrenci.bos,
                       ogrenci.puan])


    def ogrenci_cevaplari(self, ss):
        correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        current_row = 2
        #maxq = self.exam.answers[max(self.exam.answers, key=lambda a:self.exam.answers[a].question_count)].question_count
        maxq = -1
        for a in self.exam.answers:
            answer:Answers = self.exam.answers[a]
            if max(answer.mapping)+1 > maxq:
                maxq = max(answer.mapping)+1

        header = ["Öğr.No.", "Adı", "Grubu"] + [f"S{i+1}" for i in range(maxq)]
        ss.append(header)
        # set column widhts
        widhts = [90, 300, 60] + [25] * maxq
        for w, c in zip(widhts, self.get_excel_column_name()):
            ss.column_dimensions[c].width = w * 0.13
        for i,g in enumerate(self.exam.answers):
            answer:Answers = self.exam.answers[g]
            row = ["", "CEVAP", g] + [c for c in answer.answers]
            ss.append(row)
            current_row += 1
            for student in filter(lambda s:s.kitapcik == g or i == 0 and s.kitapcik in [" ", "*"], self.exam.students):
                corrects = []
                student_answers = []
                offset = 0
                for q,m in zip(range(answer.question_count), answer.mapping):
                    if q + offset != m:
                        offset += 1
                        student_answers.append(student.cevaplar[q])
                    student_answers.append(answer.get_student_answer(q, student.cevaplar))
                    corrects.append(answer.check_question(q, student.cevaplar))
                srow = [student.ogrno, student.adi, student.kitapcik] + student_answers
                ss.append(srow)
                # Fill backround colors(correct/incorrect)
                offset = 0
                for column,m in zip(range(answer.question_count), answer.mapping):
                    if column + offset != m:
                        offset += 1
                    cell = ss.cell(current_row, column+4+offset)
                    cell.fill = correct_fill if corrects[column] else incorrect_fill
                current_row += 1


    def excele_aktar(self):
        filename = f"{self.dersinAdi.text()}.xlsx"
        path = os.path.join(self.last_directory, filename)
        file_name,_ = QFileDialog.getSaveFileName(self, "Excel'e Aktar", path, "Excel dosyaları(*.xlsx)")
        if not file_name:
            return
        self.last_directory = os.path.dirname(file_name)
        wb = Workbook()

        style = NamedStyle("Baslik")
        style.font = Font(bold=True)
        style.alignment = Alignment(horizontal="center")
        wb.add_named_style(style)

        ws = wb.active
        self.default_sheet(ws, style)
        students_sheet = wb.create_sheet("Öğrenci Cevapları")
        self.ogrenci_cevaplari(students_sheet)

        if self.okIliskilendir.isChecked():
            self.exam.calculate_learnig_outcome_contributions(*self.table_model.get_group_outcome_mapping())
            outcome_sheet = wb.create_sheet("Öğrenme Çıktıları İlişkisi")
            self.iliski_yazdir(outcome_sheet)


        wb.save(file_name)

    def iliski_yazdir(self, rs):
        soru_sayisi = self.exam.answers[max(self.exam.answers, key=lambda a:self.exam.answers[a].question_count)].question_count
        widths = [100] + [50]*soru_sayisi
        for w, c in zip(widths, self.get_excel_column_name()):
            rs.column_dimensions[c].width = w * 0.13
        rs.append(["Grup Adı"]+[f"S{i+1}" for i in range(soru_sayisi)])
        for answer in self.exam.answers.values():
            alist = [answer.group] + [f"ÖK{ok+1}" for i,ok in self.exam.outcome_mapping[answer.group].items()]
            rs.append(alist)
            slist = [answer.group] + [f"{a['correct percentage']:.2f}" for a in answer.statistics]
            rs.append(slist)

        rs.append([])
        rs.append(["Grup-ÖK İlişkisi"])
        ok_sayisi = len(next(iter(self.exam.group_outcome_means.values())))
        rs.append([""]+[f"ÖK{i+1}" for i in range(ok_sayisi)])
        for answer in self.exam.group_outcome_means:
            rs.append([answer]+[f"{m:.2f}" for m in self.exam.group_outcome_means[answer].values()])

        rs.append([])
        rs.append(["Sınav-ÖK Katkısı(Grup ağırlıklı ortalaması)"])
        rs.append([""]+[f"ÖK{i+1}" for i in range(ok_sayisi)])
        rs.append([""]+[f"{k:.2f}" for k in self.exam.outcomes])


    def set_table_headers(self):
        self.notlarTablo.setColumnCount(8)
        self.notlarTablo.horizontalHeader().setStretchLastSection(True)
        self.notlarTablo.setHorizontalHeaderLabels(["Öğr.No.", "Adı", "Şube", "Kitapçık Türü", "Doğru", "Yanlış", "Boş", "Puan"])
        for i, w in enumerate([90, 300, 50, 50, 50, 50, 50, 50]):
            self.notlarTablo.setColumnWidth(i, w)

    def dosya_ac(self):
        filename,_ = QFileDialog.getOpenFileName(self, "Dosya Aç", self.last_directory, "Metin dosyaları(*.txt)")
        if not filename:
            return

        self.last_directory = os.path.dirname(filename)
        self.exam = Exam(filename, self.toplamPuan.value())
        self.dersinAdi.setText(os.path.basename(filename).removesuffix(".txt"))
        self.kitapcikSayisi.setText(str(self.exam.group_count))
        self.ogrenciSayisi.setText(str(self.exam.total_student_count))
        self.not_goster()

    def not_goster(self):
        # self.notlarTablo.clear()
        self.notlarTablo.setRowCount(len(self.exam.students))
        nt = self.notlarTablo
        if self.ogrnoSirala.isChecked():
            ogrenciler = sorted(self.exam.students, key=lambda ogr: ogr.ogrno)
        elif self.notSirala.isChecked():
            ogrenciler = sorted(self.exam.students, key=lambda ogr: ogr.puan, reverse=True)
        else:
            ogrenciler = self.exam.students

        for i, ogrenci in enumerate(ogrenciler):
            nt.setItem(i, 0, QTableWidgetItem(ogrenci.ogrno))
            nt.setItem(i, 1, QTableWidgetItem(ogrenci.adi))
            nt.setItem(i, 2, QTableWidgetItem(ogrenci.sube))
            nt.setItem(i, 3, QTableWidgetItem(ogrenci.kitapcik))
            nt.setItem(i, 4, QTableWidgetItem(str(ogrenci.dogru)))
            nt.setItem(i, 5, QTableWidgetItem(str(ogrenci.yanlis)))
            nt.setItem(i, 6, QTableWidgetItem(str(ogrenci.bos)))
            nt.setItem(i, 7, QTableWidgetItem(str(ogrenci.puan)))

    def closeEvent(self, event):
            """
            This method is called when the user closes the window.
            We use it to save our settings.
            """
            self.settings.setValue("last_directory", self.last_directory)
            super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyMainWindow()
    window.show()
    sys.exit(app.exec())
